from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

CORROSION_XLSM_TEMPLATE_PATH = (
    Path(__file__)
    .resolve()
    .with_name(
        "corrosion_entry_template.xlsm"
    )
)

CORROSION_WORKBOOK_GENERATOR_VERSION = (
    "2026-08-27-exposure-dates-v1"
)

CORROSION_METRIC_OPTIONS = [
    "penetration_rate",
    "mass_loss_rate",
    "cumulative_penetration",
    "cumulative_mass_loss",
    "maximum_pit_depth",
    "net_mass_change",
    "MCI",
    "Al-ACI",
    "ICI",

    # Retained for compatibility with observations created
    # before the more specific metric structure was introduced.
    "corrosion_rate",
]


CORROSION_UNIT_RULES = [
    # metric, reported unit, canonical unit, multiplier

    # Penetration rate
    ("penetration_rate", "µm/year", "µm/year", 1.0),
    ("penetration_rate", "µm/a", "µm/year", 1.0),
    ("penetration_rate", "mm/year", "µm/year", 1000.0),
    ("penetration_rate", "mm/a", "µm/year", 1000.0),
    ("penetration_rate", "mpy", "µm/year", 25.4),
    ("penetration_rate", "mil/year", "µm/year", 25.4),

    # Legacy corrosion-rate records are treated as penetration rates
    # only when their unit is a penetration-rate unit.
    ("corrosion_rate", "µm/year", "µm/year", 1.0),
    ("corrosion_rate", "µm/a", "µm/year", 1.0),
    ("corrosion_rate", "mm/year", "µm/year", 1000.0),
    ("corrosion_rate", "mm/a", "µm/year", 1000.0),
    ("corrosion_rate", "mpy", "µm/year", 25.4),
    ("corrosion_rate", "mil/year", "µm/year", 25.4),

    # Mass-loss rate
    ("mass_loss_rate", "g/m²/year", "g/m²/year", 1.0),
    ("mass_loss_rate", "g/m²/a", "g/m²/year", 1.0),
    ("mass_loss_rate", "g/m²·a", "g/m²/year", 1.0),
    ("mass_loss_rate", "g/m².a", "g/m²/year", 1.0),
    ("mass_loss_rate", "g/m2/a", "g/m²/year", 1.0),
    ("mass_loss_rate", "g/m2.a", "g/m²/year", 1.0),
    ("mass_loss_rate", "mg/m²/day", "g/m²/year", 0.365),
    ("mass_loss_rate", "mg/dm²/day", "g/m²/year", 36.5),
    ("mass_loss_rate", "g/dm²/month", "g/m²/year", 1200.0),
    ("mass_loss_rate", "g/dm²/year", "g/m²/year", 100.0),
    ("mass_loss_rate", "mg/cm²/day", "g/m²/year", 3650.0),

    # Cumulative penetration
    ("cumulative_penetration", "µm", "µm", 1.0),
    ("cumulative_penetration", "mm", "µm", 1000.0),
    ("cumulative_penetration", "mil", "µm", 25.4),

    # Maximum pit depth
    ("maximum_pit_depth", "µm", "µm", 1.0),
    ("maximum_pit_depth", "mm", "µm", 1000.0),
    ("maximum_pit_depth", "mil", "µm", 25.4),

    # Cumulative mass loss
    ("cumulative_mass_loss", "g/m²", "g/m²", 1.0),
    ("cumulative_mass_loss", "mg/m²", "g/m²", 0.001),
    ("cumulative_mass_loss", "g/dm²", "g/m²", 100.0),
    ("cumulative_mass_loss", "mg/dm²", "g/m²", 0.1),
    ("cumulative_mass_loss", "g/cm²", "g/m²", 10000.0),
    ("cumulative_mass_loss", "mg/cm²", "g/m²", 10.0),

    # Net mass change.
    # This can be normalized between mass/area units,
    # but must NOT be interpreted automatically as metal loss.
    ("net_mass_change", "g/m²", "g/m²", 1.0),
    ("net_mass_change", "mg/m²", "g/m²", 0.001),
    ("net_mass_change", "g/dm²", "g/m²", 100.0),
    ("net_mass_change", "mg/dm²", "g/m²", 0.1),
    ("net_mass_change", "g/cm²", "g/m²", 10000.0),
    ("net_mass_change", "mg/cm²", "g/m²", 10.0),

    # Corrosivity indices
    ("MCI", "%", "%", 1.0),
    ("MCI", "index", "index", 1.0),
    ("Al-ACI", "%", "%", 1.0),
    ("Al-ACI", "index", "index", 1.0),
    ("ICI", "%", "%", 1.0),
    ("ICI", "index", "index", 1.0),
]


DEFAULT_DENSITY_G_CM3 = {
    "Carbon steel": 7.85,
    "Mild steel": 7.85,
    "Low-alloy steel": 7.85,
    "Weathering steel": 7.85,
    "Copper-bearing steel": 7.85,

    "Iron": 7.87,
    "Ingot iron": 7.87,

    "Zinc": 7.14,
    "Copper": 8.96,

    "Aluminium": 2.70,
    "Aluminum": 2.70,

    "Lead": 11.34,
    "Nickel": 8.90,
    "Tin": 7.31,

    "Brass": 8.50,
    "Bronze": 8.80,
    "Magnesium": 1.74,
}


WORKBOOK_COLUMNS = [
    "source_code",
    "source_title",

    "site_id",
    "site_label",
    "country",

    "observation_id",

    "material",
    "exposure_period",
    "exposure_start",
    "exposure_end",
    "corrosion_metric",

    "reported_value",
    "reported_unit",

    "default_density_g_cm3",
    "density_override_g_cm3",
    "density_used_g_cm3",

    "canonical_thickness_loss_rate_um_year",
    "canonical_mass_loss_rate_g_m2_year",

    "normalization_note",

    "notes",
]

WORKBOOK_INPUT_COLUMNS = [
    "source_code",
    "source_title",
    "site_id",
    "site_label",
    "country",
    "observation_id",
    "material",
    "exposure_period",
    "exposure_start",
    "exposure_end",
    "corrosion_metric",
    "reported_value",
    "reported_unit",
    "default_density_g_cm3",
    "density_override_g_cm3",
    "density_used_g_cm3",
    "canonical_thickness_loss_rate_um_year",
    "canonical_mass_loss_rate_g_m2_year",
    "normalization_note",
    "notes",
]


CORROSION_ENTRY_REQUIRED_FIELDS = [
    "source_code",
    "site_id",
    "material",
    "corrosion_metric",
    "reported_value",
    "reported_unit",
]


CORROSION_ENTRY_FIELDS = {
    "material",
    "exposure_period",
    "exposure_start",
    "exposure_end",
    "corrosion_metric",
    "reported_value",
    "reported_unit",
    "density_override_g_cm3",
    "notes",
}


def _clean_unique(values: list[str]) -> list[str]:
    cleaned: list[str] = []
    seen: set[str] = set()

    for value in values:
        text = str(value or "").strip()

        if not text:
            continue

        key = text.casefold()

        if key in seen:
            continue

        seen.add(key)
        cleaned.append(text)

    return cleaned


def _safe_float(value: Any) -> float | None:
    if value in (None, ""):
        return None

    try:
        return float(value)
    except (TypeError, ValueError):
        return None

def _normalize_unit_text(value: Any) -> str:
    text = str(value or "").strip()

    replacements = {
        "μ": "µ",
        "²": "²",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    return text

def parse_exposure_years(
    exposure_period: str,
) -> float | None:
    """
    Convert a simple exposure-duration string to years.

    Examples:
        1 year      -> 1.0
        2 years     -> 2.0
        6 months    -> 0.5
        90 days     -> 90 / 365.25
        18 months   -> 1.5
        0.5 year    -> 0.5
        3 weeks     -> 21 / 365.25

    Ambiguous ranges or unrecognised text return None.
    """

    import re

    text = str(
        exposure_period or ""
    ).strip().lower()

    if not text:
        return None

    text = (
        text
        .replace("–", "-")
        .replace("—", "-")
    )

    # Do not guess ranges such as 1-2 years.
    if re.search(
        r"\d\s*-\s*\d",
        text,
    ):
        return None

    match = re.fullmatch(
        r"\s*"
        r"(\d+(?:\.\d+)?)"
        r"\s*"
        r"(years?|yrs?|yr|y|"
        r"months?|mos?|mo|"
        r"weeks?|wks?|wk|"
        r"days?|d)"
        r"\s*",
        text,
    )

    if match is None:
        return None

    value = float(
        match.group(1)
    )

    unit = match.group(2)

    if value <= 0:
        return None

    if unit in {
        "year",
        "years",
        "yr",
        "yrs",
        "y",
    }:
        return value

    if unit in {
        "month",
        "months",
        "mo",
        "mos",
    }:
        return value / 12.0

    if unit in {
        "week",
        "weeks",
        "wk",
        "wks",
    }:
        return (
            value * 7.0 / 365.25
        )

    if unit in {
        "day",
        "days",
        "d",
    }:
        return (
            value / 365.25
        )

    return None

def is_valid_partial_iso_date(
    value: Any,
) -> bool:
    """
    Accept source-supported calendar precision:

        YYYY
        YYYY-MM
        YYYY-MM-DD

    Blank values are allowed.

    This validates the stored canonical text only.
    Natural-language Excel input is normalized separately by VBA.
    """

    import re
    from datetime import date

    text = str(
        value or ""
    ).strip()

    if not text:
        return True

    match = re.fullmatch(
        r"(\d{4})"
        r"(?:-(\d{2})"
        r"(?:-(\d{2}))?"
        r")?",
        text,
    )

    if match is None:
        return False

    year = int(
        match.group(1)
    )

    if not 1 <= year <= 9999:
        return False

    month_text = match.group(2)
    day_text = match.group(3)

    if month_text is None:
        return True

    month = int(
        month_text
    )

    if not 1 <= month <= 12:
        return False

    if day_text is None:
        return True

    day = int(
        day_text
    )

    try:
        date(
            year,
            month,
            day,
        )
    except ValueError:
        return False

    return True

def _get_unit_rule(
    metric: str,
    reported_unit: str,
) -> tuple[str, float] | None:

    clean_metric = str(metric or "").strip()
    clean_unit = _normalize_unit_text(
        reported_unit
    )

    for (
        rule_metric,
        rule_unit,
        canonical_unit,
        multiplier,
    ) in CORROSION_UNIT_RULES:

        if (
            rule_metric == clean_metric
            and _normalize_unit_text(
                rule_unit
            ) == clean_unit
        ):
            return (
                canonical_unit,
                float(multiplier),
            )

    return None


def get_default_density_g_cm3(
    material: str,
) -> float | None:

    clean_material = str(
        material or ""
    ).strip()

    if not clean_material:
        return None

    # Exact match first.
    if clean_material in DEFAULT_DENSITY_G_CM3:
        return float(
            DEFAULT_DENSITY_G_CM3[
                clean_material
            ]
        )

    # Case-insensitive fallback.
    material_key = clean_material.casefold()

    for (
        known_material,
        density,
    ) in DEFAULT_DENSITY_G_CM3.items():

        if (
            known_material.casefold()
            == material_key
        ):
            return float(density)

    return None

def normalize_corrosion_observation(
    *,
    material: str,
    exposure_period: str,
    corrosion_metric: str,
    reported_value: float,
    reported_unit: str,
    density_override_g_cm3: float | None = None,
) -> dict[str, Any]:
    """
    Convert a source-reported corrosion observation into up to
    two canonical general-corrosion rates:

        thickness-loss rate = µm/year
        mass-loss rate      = g/m²/year

    The original reported metric/value/unit remain authoritative.

    Where density is available, thickness-loss and mass-loss
    representations are cross-converted.

    Cumulative quantities additionally require a usable exposure
    duration before they can be annualized.

    Metrics which are not measures of general corrosion
    (maximum pit depth, net mass change, MCI, Al-ACI, ICI)
    deliberately receive neither canonical rate.
    """

    material = str(
        material or ""
    ).strip()

    exposure_period = str(
        exposure_period or ""
    ).strip()

    corrosion_metric = str(
        corrosion_metric or ""
    ).strip()

    reported_unit = (
        _normalize_unit_text(
            reported_unit
        )
    )

    reported_value = float(
        reported_value
    )

    if (
        corrosion_metric != "net_mass_change"
        and reported_value < 0
    ):
        raise ValueError(
            "Negative reported values are only accepted "
            "for net_mass_change."
        )

    default_density = (
        get_default_density_g_cm3(
            material
        )
    )

    density_override = (
        _safe_float(
            density_override_g_cm3
        )
    )

    if (
        density_override is not None
        and density_override <= 0
    ):
        raise ValueError(
            "density_override_g_cm3 must be greater than zero."
        )

    if density_override is not None:
        density_used = density_override
        density_basis = "curator_override"

    elif default_density is not None:
        density_used = default_density
        density_basis = "default_material_density"

    else:
        density_used = None
        density_basis = ""

    exposure_years = (
        parse_exposure_years(
            exposure_period
        )
    )

    canonical_thickness_loss_rate = None
    canonical_mass_loss_rate = None

    normalization_note = ""

    # -----------------------------------------------------
    # 1. Reported penetration/thickness-loss rate
    # -----------------------------------------------------

    if corrosion_metric in {
        "penetration_rate",
        "corrosion_rate",
    }:
        unit_rule = _get_unit_rule(
            corrosion_metric,
            reported_unit,
        )

        if unit_rule is None:
            raise ValueError(
                "Unsupported metric/unit combination: "
                f"{corrosion_metric} + {reported_unit}"
            )

        canonical_unit, multiplier = unit_rule

        if canonical_unit != "µm/year":
            raise ValueError(
                "Penetration-rate conversion must "
                "first produce µm/year."
            )

        canonical_thickness_loss_rate = (
            reported_value
            * multiplier
        )

        if density_used is not None:
            canonical_mass_loss_rate = (
                canonical_thickness_loss_rate
                * density_used
            )

            normalization_note = (
                "Converted reported penetration rate to "
                "µm/year and derived mass-loss rate using "
                f"density {density_used:g} g/cm³."
            )

        else:
            normalization_note = (
                "Converted reported penetration rate to "
                "µm/year. Mass-loss rate could not be "
                "derived because density is unavailable."
            )

    # -----------------------------------------------------
    # 2. Reported mass-loss rate
    # -----------------------------------------------------

    elif corrosion_metric == "mass_loss_rate":
        unit_rule = _get_unit_rule(
            corrosion_metric,
            reported_unit,
        )

        if unit_rule is None:
            raise ValueError(
                "Unsupported metric/unit combination: "
                f"{corrosion_metric} + {reported_unit}"
            )

        canonical_unit, multiplier = unit_rule

        if canonical_unit != "g/m²/year":
            raise ValueError(
                "Mass-loss-rate conversion must "
                "first produce g/m²/year."
            )

        canonical_mass_loss_rate = (
            reported_value
            * multiplier
        )

        if density_used is not None:
            canonical_thickness_loss_rate = (
                canonical_mass_loss_rate
                / density_used
            )

            normalization_note = (
                "Converted reported mass-loss rate to "
                "g/m²/year and derived thickness-loss rate "
                f"using density {density_used:g} g/cm³."
            )

        else:
            normalization_note = (
                "Converted reported mass-loss rate to "
                "g/m²/year. Thickness-loss rate could not "
                "be derived because density is unavailable."
            )

    # -----------------------------------------------------
    # 3. Reported cumulative penetration
    # -----------------------------------------------------

    elif corrosion_metric == "cumulative_penetration":
        unit_rule = _get_unit_rule(
            corrosion_metric,
            reported_unit,
        )

        if unit_rule is None:
            raise ValueError(
                "Unsupported metric/unit combination: "
                f"{corrosion_metric} + {reported_unit}"
            )

        canonical_unit, multiplier = unit_rule

        if canonical_unit != "µm":
            raise ValueError(
                "Cumulative penetration must first "
                "normalize to µm."
            )

        cumulative_penetration_um = (
            reported_value
            * multiplier
        )

        if exposure_years is None:

            if not exposure_period:
                normalization_note = (
                    "Exposure period not reported. "
                    "Cumulative penetration is preserved as "
                    "reported but cannot be converted to annual "
                    "canonical corrosion rates."
                )

            else:
                normalization_note = (
                    "Cumulative penetration is preserved as "
                    "reported but cannot be converted to annual "
                    "canonical corrosion rates because the "
                    "exposure duration could not be interpreted."
                )

        else:
            canonical_thickness_loss_rate = (
                cumulative_penetration_um
                / exposure_years
            )

            if density_used is not None:
                canonical_mass_loss_rate = (
                    canonical_thickness_loss_rate
                    * density_used
                )

                normalization_note = (
                    "Converted cumulative penetration to "
                    "average annual thickness-loss rate over "
                    f"{exposure_years:g} year(s), then derived "
                    "mass-loss rate using density "
                    f"{density_used:g} g/cm³."
                )

            else:
                normalization_note = (
                    "Converted cumulative penetration to "
                    "average annual thickness-loss rate over "
                    f"{exposure_years:g} year(s). Mass-loss "
                    "rate could not be derived because density "
                    "is unavailable."
                )

    # -----------------------------------------------------
    # 4. Reported cumulative mass loss
    # -----------------------------------------------------

    elif corrosion_metric == "cumulative_mass_loss":
        unit_rule = _get_unit_rule(
            corrosion_metric,
            reported_unit,
        )

        if unit_rule is None:
            raise ValueError(
                "Unsupported metric/unit combination: "
                f"{corrosion_metric} + {reported_unit}"
            )

        canonical_unit, multiplier = unit_rule

        if canonical_unit != "g/m²":
            raise ValueError(
                "Cumulative mass loss must first "
                "normalize to g/m²."
            )

        cumulative_mass_loss_g_m2 = (
            reported_value
            * multiplier
        )

        if exposure_years is None:

            if not exposure_period:
                normalization_note = (
                    "Exposure period not reported. "
                    "Cumulative mass loss is preserved as "
                    "reported but cannot be converted to annual "
                    "canonical corrosion rates."
                )

            else:
                normalization_note = (
                    "Cumulative mass loss is preserved as "
                    "reported but cannot be converted to annual "
                    "canonical corrosion rates because the "
                    "exposure duration could not be interpreted."
                )

        else:
            canonical_mass_loss_rate = (
                cumulative_mass_loss_g_m2
                / exposure_years
            )

            if density_used is not None:
                canonical_thickness_loss_rate = (
                    canonical_mass_loss_rate
                    / density_used
                )

                normalization_note = (
                    "Converted cumulative mass loss to "
                    "average annual mass-loss rate over "
                    f"{exposure_years:g} year(s), then derived "
                    "thickness-loss rate using density "
                    f"{density_used:g} g/cm³."
                )

            else:
                normalization_note = (
                    "Converted cumulative mass loss to "
                    "average annual mass-loss rate over "
                    f"{exposure_years:g} year(s). Thickness-loss "
                    "rate could not be derived because density "
                    "is unavailable."
                )

    # -----------------------------------------------------
    # 5. Metrics intentionally not mapped to general
    #    corrosion rates
    # -----------------------------------------------------

    elif corrosion_metric == "maximum_pit_depth":
        if _get_unit_rule(
            corrosion_metric,
            reported_unit,
        ) is None:
            raise ValueError(
                "Unsupported metric/unit combination: "
                f"{corrosion_metric} + {reported_unit}"
            )

        normalization_note = (
            "Maximum pit depth is a localized-corrosion "
            "metric and is not converted to the canonical "
            "general thickness-loss or mass-loss rates."
        )

    elif corrosion_metric == "net_mass_change":
        if _get_unit_rule(
            corrosion_metric,
            reported_unit,
        ) is None:
            raise ValueError(
                "Unsupported metric/unit combination: "
                f"{corrosion_metric} + {reported_unit}"
            )

        normalization_note = (
            "Net mass change is not assumed to equal cleaned "
            "metal loss and therefore is not converted to "
            "canonical corrosion rates."
        )

    elif corrosion_metric in {
        "MCI",
        "Al-ACI",
        "ICI",
    }:
        if _get_unit_rule(
            corrosion_metric,
            reported_unit,
        ) is None:
            raise ValueError(
                "Unsupported metric/unit combination: "
                f"{corrosion_metric} + {reported_unit}"
            )

        normalization_note = (
            f"{corrosion_metric} is a corrosivity index and "
            "cannot be converted to canonical general "
            "corrosion rates."
        )

    else:
        raise ValueError(
            "Unsupported corrosion metric: "
            f"{corrosion_metric}"
        )

    return {
        "canonical_thickness_loss_rate_um_year": (
            canonical_thickness_loss_rate
        ),

        "canonical_mass_loss_rate_g_m2_year": (
            canonical_mass_loss_rate
        ),

        "default_density_g_cm3": (
            default_density
        ),

        "density_override_g_cm3": (
            density_override
        ),

        "density_g_cm3": (
            density_used
        ),

        "density_basis": (
            density_basis
        ),

        "normalization_note": (
            normalization_note
        ),

        "exposure_years": (
            exposure_years
        ),

        # Temporary backward-compatibility fields.
        # The current database still knows normalized_value.
        "normalized_value": (
            canonical_thickness_loss_rate
        ),

        "normalized_unit": (
            "µm/year"
            if canonical_thickness_loss_rate is not None
            else ""
        ),
    }

def build_corrosion_entry_workbook(
    *,
    source_row: dict[str, Any],
    site_links: list[dict[str, Any]],
    site_lookup: dict[str, dict[str, Any]],
    existing_observations: list[dict[str, Any]],
    metal_options: list[str],
    exposure_options: list[str],
    blank_rows_per_site: int = 8,
    macro_template_path: (
        str | Path | None
    ) = None,
) -> bytes:
    """
    Build a source-first Excel workbook for corrosion observation entry.

    Structure:
        source
            -> linked sites
                -> existing observations
                -> blank observation-entry rows

    Source and site values are repeated on every row intentionally.
    Merged cells are avoided because they make later Excel import fragile.
    """

    source_code = str(
        source_row.get("source_code", "") or ""
    ).strip()

    source_title = str(
        source_row.get("source_title", "") or ""
    ).strip()

    macro_enabled = (
        macro_template_path
        is not None
    )    

    if not source_code:
        raise ValueError(
            "A source_code is required to generate a corrosion workbook."
        )

    if macro_enabled:
        # The macro-enabled workbook only needs one starter
        # row per linked site. Additional observations are
        # created directly inside Excel using the + control.
        blank_rows_per_site = 1

    else:
        blank_rows_per_site = max(
            1,
            min(
                int(
                    blank_rows_per_site
                ),
                50,
            ),
        )

    # ---------------------------------------------------------
    # Identify unique sites linked to this source
    # ---------------------------------------------------------

    unique_links_by_site: dict[str, dict[str, Any]] = {}

    for link in site_links:
        site_id = str(
            link.get("site_id", "") or ""
        ).strip()

        if site_id and site_id not in unique_links_by_site:
            unique_links_by_site[site_id] = link

    if not unique_links_by_site:
        raise ValueError(
            f"Source {source_code} has no site-source links. "
            "Link the source to site(s) first."
        )

    # ---------------------------------------------------------
    # Group existing observations by site
    # ---------------------------------------------------------

    observations_by_site: dict[
        str,
        list[dict[str, Any]]
    ] = {}

    for observation in existing_observations:
        observation_source = str(
            observation.get("source_code", "") or ""
        ).strip().casefold()

        if observation_source != source_code.casefold():
            continue

        site_id = str(
            observation.get("site_id", "") or ""
        ).strip()

        # Observation worksheets are based on curated
        # site-source relationships.
        if site_id not in unique_links_by_site:
            continue

        observations_by_site.setdefault(
            site_id,
            [],
        ).append(observation)

    for site_observations in observations_by_site.values():
        site_observations.sort(
            key=lambda row: (
                str(
                    row.get("material", "") or ""
                ).casefold(),
                str(
                    row.get("exposure_period", "") or ""
                ).casefold(),
                str(
                    row.get(
                        "exposure_start",
                        "",
                    ) or ""
                ).casefold(),
                str(
                    row.get(
                        "exposure_end",
                        "",
                    ) or ""
                ).casefold(),
                str(
                    row.get("corrosion_metric", "") or ""
                ).casefold(),
                int(row.get("id", 0) or 0),
            )
        )

    # ---------------------------------------------------------
    # Dropdown lists
    # ---------------------------------------------------------

    metal_options = _clean_unique(
        list(metal_options)
        + [
            str(
                observation.get("material", "") or ""
            )
            for observation in existing_observations
        ]
        + list(DEFAULT_DENSITY_G_CM3.keys())
    )

    exposure_options = _clean_unique(
        list(exposure_options)
        + [
            str(
                observation.get(
                    "exposure_period",
                    "",
                ) or ""
            )
            for observation in existing_observations
        ]
    )

    metric_options = _clean_unique(
        CORROSION_METRIC_OPTIONS
        + [
            str(
                observation.get(
                    "corrosion_metric",
                    "",
                ) or ""
            )
            for observation in existing_observations
        ]
    )

    unit_options = _clean_unique(
        [
            unit
            for _, unit, _, _
            in CORROSION_UNIT_RULES
        ]
        + [
            str(
                observation.get("unit", "") or ""
            )
            for observation in existing_observations
        ]
    )

    # ---------------------------------------------------------
    # Workbook
    # ---------------------------------------------------------

    if macro_enabled:

        template_path = Path(
            macro_template_path
        )

        if not template_path.exists():
            raise FileNotFoundError(
                "Macro-enabled corrosion workbook "
                f"template not found: {template_path}"
            )

        workbook = load_workbook(
            template_path,
            keep_vba=True,
        )

        if (
            "Corrosion Observations"
            not in workbook.sheetnames
        ):
            raise ValueError(
                "The XLSM template must contain a "
                "`Corrosion Observations` worksheet."
            )

        sheet = workbook[
            "Corrosion Observations"
        ]

        # Keep the worksheet object itself because the
        # VBA Worksheet_SelectionChange event belongs to it.
        # Only clear its spreadsheet contents.

        if sheet.max_row > 0:
            sheet.delete_rows(
                1,
                sheet.max_row,
            )

        # Guide and Lists contain no VBA event code,
        # so they can safely be regenerated.

        for disposable_sheet_name in [
            "Guide",
            "Lists",
        ]:
            if (
                disposable_sheet_name
                in workbook.sheetnames
            ):
                workbook.remove(
                    workbook[
                        disposable_sheet_name
                    ]
                )

        guide_sheet = (
            workbook.create_sheet(
                "Guide"
            )
        )

        lists_sheet = (
            workbook.create_sheet(
                "Lists"
            )
        )

    else:

        workbook = Workbook()

        sheet = workbook.active
        sheet.title = (
            "Corrosion Observations"
        )

        guide_sheet = (
            workbook.create_sheet(
                "Guide"
            )
        )

        lists_sheet = (
            workbook.create_sheet(
                "Lists"
            )
        )

    # =========================================================
    # User guide sheet
    # =========================================================

    guide_sheet.sheet_view.showGridLines = False
    guide_sheet.sheet_properties.tabColor = "5B9BD5"

    guide_sheet["A1"] = (
        "Corrosion Atlas — Observation Entry Guide"
    )

    guide_sheet["A1"].font = Font(
        bold=True,
        size=16,
        color="FFFFFF",
    )

    guide_sheet["A1"].fill = PatternFill(
        "solid",
        fgColor="17365D",
    )

    guide_sheet["A1"].alignment = Alignment(
        vertical="center",
    )

    guide_sheet.merge_cells(
        "A1:F1"
    )

    guide_sheet.row_dimensions[1].height = 28

    guide_sheet["A3"] = "Workbook purpose"
    guide_sheet["A3"].font = Font(
        bold=True,
        size=12,
        color="17365D",
    )

    guide_sheet["A4"] = (
        "Enter corrosion observations exactly as reported "
        "by the selected source. Calculated fields are only "
        "a workbook preview; the curator app independently "
        "validates and recalculates them during import."
    )

    guide_sheet.merge_cells(
        "A4:F5"
    )

    guide_sheet["A4"].alignment = Alignment(
        wrap_text=True,
        vertical="top",
    )

    guide_sheet["A7"] = "Cell colours"
    guide_sheet["A7"].font = Font(
        bold=True,
        size=12,
        color="17365D",
    )

    guide_rows = [
        (
            "A9",
            "Blue",
            "Source/site context — normally do not edit.",
            "D9EAF7",
        ),
        (
            "A10",
            "Yellow",
            "Curator input — enter or select source data here.",
            "FFF2CC",
        ),
        (
            "A11",
            "Grey",
            "Calculated field — generated automatically.",
            "E7E6E6",
        ),
        (
            "A12",
            "Green",
            "Existing database observation.",
            "E2F0D9",
        ),
        (
            "A13",
            "Red/Pink",
            "Incomplete observation row — required entry fields are missing.",
            "F4CCCC",
        ),
    ]

    for (
        cell_ref,
        label,
        description,
        fill_color,
    ) in guide_rows:
        guide_sheet[cell_ref] = label
        guide_sheet[cell_ref].fill = PatternFill(
            "solid",
            fgColor=fill_color,
        )
        guide_sheet[cell_ref].font = Font(
            bold=True,
        )

        description_cell = (
            f"B{guide_sheet[cell_ref].row}"
        )

        guide_sheet[description_cell] = description

    guide_sheet["A16"] = "Core input fields"
    guide_sheet["A16"].font = Font(
        bold=True,
        size=12,
        color="17365D",
    )

    guide_sheet["A18"] = "material"
    guide_sheet["B18"] = (
        "Metal/material exposed at the site."
    )

    guide_sheet["A19"] = "exposure_period"
    guide_sheet["B19"] = (
        "Optional exposure duration, e.g. 90 days, 6 months, "
        "1 year, 16 years. Leave blank when the source does not "
        "report a duration. Cumulative quantities cannot be "
        "annualized without a known duration."
    )

    guide_sheet["A20"] = "exposure_start"
    guide_sheet["B20"] = (
        "Optional calendar start of exposure. Use YYYY, YYYY-MM, "
        "or YYYY-MM-DD according to the precision actually reported "
        "by the source. Do not invent missing months or days."
    )

    guide_sheet["A21"] = "exposure_end"
    guide_sheet["B21"] = (
        "Optional calendar end of exposure. Use YYYY, YYYY-MM, "
        "or YYYY-MM-DD according to the precision actually reported "
        "by the source. Do not invent missing months or days."
    )

    guide_sheet["A22"] = "corrosion_metric"
    guide_sheet["B22"] = (
        "Physical quantity actually reported by the source."
    )

    guide_sheet["A23"] = "reported_value"
    guide_sheet["B23"] = (
        "Numerical value exactly as reported."
    )

    guide_sheet["A24"] = "reported_unit"
    guide_sheet["B24"] = (
        "Original unit used by the source."
    )

    guide_sheet["A25"] = "Main corrosion metrics"
    guide_sheet["A25"].font = Font(
        bold=True,
        size=12,
        color="17365D",
    )

    metric_guide = [
        (
            "penetration_rate",
            "Thickness/penetration loss already expressed as a rate.",
        ),
        (
            "mass_loss_rate",
            "Mass loss per exposed area per unit time.",
        ),
        (
            "cumulative_penetration",
            "Total thickness loss after the stated exposure duration.",
        ),
        (
            "cumulative_mass_loss",
            "Total mass loss after the stated exposure duration.",
        ),
        (
            "maximum_pit_depth",
            "Localized pit depth; not converted to general corrosion rate.",
        ),
        (
            "net_mass_change",
            "Net specimen mass change; not assumed to equal cleaned metal loss.",
        ),
        (
            "MCI / Al-ACI / ICI",
            "Corrosivity indices; not converted to general corrosion rate.",
        ),
    ]

    for offset, (
        metric_name,
        metric_description,
    ) in enumerate(
        metric_guide,
        start=27,
    ):
        guide_sheet.cell(
            row=offset,
            column=1,
            value=metric_name,
        ).font = Font(
            bold=True,
        )

        guide_sheet.cell(
            row=offset,
            column=2,
            value=metric_description,
        )

    guide_sheet["A36"] = "Canonical outputs"
    guide_sheet["A36"].font = Font(
        bold=True,
        size=12,
        color="17365D",
    )

    guide_sheet["A38"] = (
        "Thickness-loss rate"
    )
    guide_sheet["B38"] = (
        "µm/year"
    )

    guide_sheet["A39"] = (
        "Mass-loss rate"
    )
    guide_sheet["B39"] = (
        "g/m²/year"
    )

    guide_sheet["A41"] = (
        "When density and exposure duration permit, "
        "both canonical rates are calculated from the "
        "same reported observation."
    )

    guide_sheet["A43"] = (
        "Generator build"
    )

    guide_sheet["A43"].font = Font(
        bold=True,
        color="17365D",
    )

    guide_sheet["B43"] = (
        CORROSION_WORKBOOK_GENERATOR_VERSION
    )    

    if macro_enabled:

        guide_sheet["A45"] = (
            "Adding another observation"
        )

        guide_sheet["A45"].font = Font(
            bold=True,
            size=12,
            color="17365D",
        )

        guide_sheet["A47"] = "＋"

        guide_sheet["A47"].fill = PatternFill(
            "solid",
            fgColor="E2F0D9",
        )

        guide_sheet["A47"].font = Font(
            bold=True,
            size=14,
            color="548235",
        )

        guide_sheet["B47"] = (
            "Click the green + button beside any observation "
            "to create a new observation immediately underneath. "
            "The new row keeps the source/site context and "
            "dropdown selections, but observation_id and "
            "reported_value are cleared automatically."
        )

    guide_sheet.merge_cells(
        "A41:F42"
    )

    guide_sheet["A41"].alignment = Alignment(
        wrap_text=True,
        vertical="top",
    )

    guide_sheet.column_dimensions["A"].width = 30
    guide_sheet.column_dimensions["B"].width = 68

    for column_letter in [
        "C",
        "D",
        "E",
        "F",
    ]:
        guide_sheet.column_dimensions[
            column_letter
        ].width = 12    

    # =========================================================
    # Hidden lookup sheet
    # =========================================================

    lists_sheet.append(
        [
            "metric_unit_key",
            "normalized_unit",
            "multiplier",
        ]
    )

    for (
        metric,
        unit,
        normalized_unit,
        multiplier,
    ) in CORROSION_UNIT_RULES:
        lists_sheet.append(
            [
                f"{metric}|{unit}",
                normalized_unit,
                multiplier,
            ]
        )

    unit_rule_end_row = lists_sheet.max_row

    # Density table

    density_start_row = 2

    lists_sheet["E1"] = "material"
    lists_sheet["F1"] = "density_g_cm3"

    for row_number, (
        material,
        density,
    ) in enumerate(
        DEFAULT_DENSITY_G_CM3.items(),
        start=density_start_row,
    ):
        lists_sheet.cell(
            row=row_number,
            column=5,
            value=material,
        )

        lists_sheet.cell(
            row=row_number,
            column=6,
            value=density,
        )

    density_end_row = (
        density_start_row
        + len(DEFAULT_DENSITY_G_CM3)
        - 1
    )

    # Metal dropdown

    lists_sheet["H1"] = "metal_options"

    for index, value in enumerate(
        metal_options,
        start=2,
    ):
        lists_sheet.cell(
            row=index,
            column=8,
            value=value,
        )

    metal_end_row = max(
        2,
        len(metal_options) + 1,
    )

    # Exposure dropdown

    lists_sheet["I1"] = "exposure_options"

    for index, value in enumerate(
        exposure_options,
        start=2,
    ):
        lists_sheet.cell(
            row=index,
            column=9,
            value=value,
        )

    exposure_end_row = max(
        2,
        len(exposure_options) + 1,
    )

    # Metric dropdown

    lists_sheet["J1"] = "metric_options"

    for index, value in enumerate(
        metric_options,
        start=2,
    ):
        lists_sheet.cell(
            row=index,
            column=10,
            value=value,
        )

    metric_end_row = max(
        2,
        len(metric_options) + 1,
    )

    # Unit dropdown

    lists_sheet["K1"] = "unit_options"

    for index, value in enumerate(
        unit_options,
        start=2,
    ):
        lists_sheet.cell(
            row=index,
            column=11,
            value=value,
        )

    unit_end_row = max(
        2,
        len(unit_options) + 1,
    )

    # ---------------------------------------------------------
    # Metric-specific unit dropdown lists
    # ---------------------------------------------------------

    lists_sheet["L1"] = "blank_unit_option"
    lists_sheet["L2"] = ""

    metric_unit_ranges: dict[
        str,
        tuple[str, int, int]
    ] = {}

    next_unit_column = 13  # Column M

    for metric_name in (
        CORROSION_METRIC_OPTIONS
    ):
        valid_units = _clean_unique(
            [
                unit
                for (
                    rule_metric,
                    unit,
                    _,
                    _,
                )
                in CORROSION_UNIT_RULES
                if rule_metric == metric_name
            ]
        )

        if not valid_units:
            continue

        column_index = (
            next_unit_column
        )

        column_letter = (
            get_column_letter(
                column_index
            )
        )

        safe_metric_name = (
            metric_name
            .replace("-", "_")
        )

        range_name = (
            "CorrosionUnits_"
            f"{safe_metric_name}"
        )

        lists_sheet.cell(
            row=1,
            column=column_index,
            value=(
                f"{metric_name}_units"
            ),
        )

        for row_offset, unit_value in enumerate(
            valid_units,
            start=2,
        ):
            lists_sheet.cell(
                row=row_offset,
                column=column_index,
                value=unit_value,
            )

        end_row = (
            len(valid_units) + 1
        )

        metric_unit_ranges[
            metric_name
        ] = (
            range_name,
            column_index,
            end_row,
        )

        next_unit_column += 1

    # Excel data validation does not reliably accept
    # direct references to a different worksheet.
    # Named ranges are therefore used.

    lists_ref = quote_sheetname(
        lists_sheet.title
    )

    workbook.defined_names.add(
        DefinedName(
            "CorrosionMetalOptions",
            attr_text=(
                f"{lists_ref}!$H$2:$H${metal_end_row}"
            ),
        )
    )

    workbook.defined_names.add(
        DefinedName(
            "CorrosionExposureOptions",
            attr_text=(
                f"{lists_ref}!$I$2:$I${exposure_end_row}"
            ),
        )
    )

    workbook.defined_names.add(
        DefinedName(
            "CorrosionMetricOptions",
            attr_text=(
                f"{lists_ref}!$J$2:$J${metric_end_row}"
            ),
        )
    )

    workbook.defined_names.add(
        DefinedName(
            "CorrosionUnitOptions",
            attr_text=(
                f"{lists_ref}!$K$2:$K${unit_end_row}"
            ),
        )
    )

    workbook.defined_names.add(
        DefinedName(
            "CorrosionUnits_blank",
            attr_text=(
                f"{lists_ref}!$L$2:$L$2"
            ),
        )
    )

    for (
        metric_name,
        (
            range_name,
            column_index,
            end_row,
        ),
    ) in metric_unit_ranges.items():

        column_letter = (
            get_column_letter(
                column_index
            )
        )

        workbook.defined_names.add(
            DefinedName(
                range_name,
                attr_text=(
                    f"{lists_ref}!"
                    f"${column_letter}$2:"
                    f"${column_letter}${end_row}"
                ),
            )
        )

    lists_sheet.sheet_state = "hidden"

    # =========================================================
    # Main observation sheet
    # =========================================================

    sheet.append(WORKBOOK_COLUMNS)

    # Hidden calculation helpers.
    # These avoid relying on newer Excel functions such as LET.
    sheet["V1"] = "__unit_factor"
    sheet["W1"] = "__exposure_years"

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    header_border = Border(
        bottom=Side(
            style="thin",
            color="FFFFFF",
        )
    )

    thin_gray = Side(
        style="thin",
        color="D9E1F2",
    )

    context_header_fill = PatternFill(
        "solid",
        fgColor="4472C4",
    )

    input_header_fill = PatternFill(
        "solid",
        fgColor="BF9000",
    )

    density_header_fill = PatternFill(
        "solid",
        fgColor="7F8C8D",
    )

    output_header_fill = PatternFill(
        "solid",
        fgColor="548235",
    )

    note_header_fill = PatternFill(
        "solid",
        fgColor="7030A0",
    )

    for column_index, cell in enumerate(
        sheet[1],
        start=1,
    ):

        if 1 <= column_index <= 6:
            cell.fill = context_header_fill

        elif 7 <= column_index <= 13:
            cell.fill = input_header_fill

        elif 14 <= column_index <= 16:
            cell.fill = density_header_fill

        elif 17 <= column_index <= 18:
            cell.fill = output_header_fill

        else:
            cell.fill = note_header_fill

        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        cell.border = header_border

    sheet.row_dimensions[1].height = 46

    # Freeze only the header row.
    # Do not freeze any columns; this keeps horizontal
    # navigation predictable.
    sheet.freeze_panes = "A2"

    # ---------------------------------------------------------
    # Header explanations
    # ---------------------------------------------------------

    header_comments = {
        "observation_id": (
            "Existing observations carry a database ID. "
            "Leave blank for new observations."
        ),

        "material": (
            "Choose a metal/material already known to "
            "the curator database."
        ),

        "exposure_period": (
            "Optional. Enter the exposure duration exactly as "
            "reported by the source. Leave blank when the source "
            "does not report a duration. A duration is required "
            "only to annualize cumulative corrosion quantities."
        ),

        "exposure_start": (
            "Optional calendar start of exposure. "
            "Stored as text using YYYY, YYYY-MM, or YYYY-MM-DD. "
            "Use only the precision explicitly reported by the source."
        ),

        "exposure_end": (
            "Optional calendar end of exposure. "
            "Stored as text using YYYY, YYYY-MM, or YYYY-MM-DD. "
            "Use only the precision explicitly reported by the source."
        ),

        "corrosion_metric": (
            "Select the physical quantity actually "
            "reported by the source."
        ),

        "reported_value": (
            "Enter the numerical value exactly as "
            "reported by the source."
        ),

        "reported_unit": (
            "Choose the unit actually reported by the "
            "source. Only approved units are accepted "
            "during import."
        ),

        "default_density_g_cm3": (
            "Automatically looked up from the material "
            "when a generic density is available."
        ),

        "density_override_g_cm3": (
            "Optional. Use only when a source-specific "
            "or alloy-specific density should replace "
            "the default."
        ),

        "density_used_g_cm3": (
            "Calculated from the override when supplied; "
            "otherwise from the default density."
        ),

        "canonical_thickness_loss_rate_um_year": (
            "Canonical general corrosion thickness-loss rate "
            "in µm/year. Calculated from the reported value, "
            "exposure duration and density where required. "
            "The curator app recalculates this on import."
        ),

        "canonical_mass_loss_rate_g_m2_year": (
            "Canonical general corrosion mass-loss rate in "
            "g/m²/year. Calculated from the reported value, "
            "exposure duration and density where required. "
            "The curator app recalculates this on import."
        ),

        "normalization_note": (
            "Explains how the reported observation was "
            "converted to the canonical corrosion rate, "
            "or why conversion is not valid."
        ),
    }

    for (
        column_name,
        comment_text,
    ) in header_comments.items():

        column_index = (
            WORKBOOK_COLUMNS.index(
                column_name
            )
            + 1
        )

        sheet.cell(
            row=1,
            column=column_index,
        ).comment = Comment(
            comment_text,
            "Corrosion Atlas",
        )

    # ---------------------------------------------------------
    # Source -> sites -> observations
    # ---------------------------------------------------------

    site_ids = sorted(
        unique_links_by_site.keys(),
        key=lambda site_id: (
            str(
                unique_links_by_site[
                    site_id
                ].get(
                    "site_label",
                    "",
                ) or ""
            ).casefold(),
            site_id.casefold(),
        ),
    )

    first_data_row = 2

    first_row_by_site: dict[
        str,
        int
    ] = {}    

    for site_id in site_ids:
        link = unique_links_by_site[
            site_id
        ]

        site = site_lookup.get(
            site_id,
            {},
        )

        site_label = str(
            site.get("site_label")
            or link.get("site_label")
            or ""
        ).strip()

        country = str(
            site.get(
                "modern_country_location",
                "",
            ) or ""
        ).strip()

        site_existing = (
            observations_by_site.get(
                site_id,
                [],
            )
        )

        rows_for_site: list[
            dict[str, Any]
        ] = []

        # Existing records first

        for observation in site_existing:
            rows_for_site.append(
                {
                    "source_code": source_code,
                    "source_title": source_title,

                    "site_id": site_id,
                    "site_label": site_label,
                    "country": country,

                    "observation_id": (
                        observation.get(
                            "id",
                            "",
                        )
                    ),

                    "material": (
                        observation.get(
                            "material",
                            "",
                        )
                    ),

                    "exposure_period": (
                        observation.get(
                            "exposure_period",
                            "",
                        )
                    ),

                    "exposure_start": (
                        observation.get(
                            "exposure_start",
                            "",
                        )
                    ),

                    "exposure_end": (
                        observation.get(
                            "exposure_end",
                            "",
                        )
                    ),

                    "corrosion_metric": (
                        observation.get(
                            "corrosion_metric",
                            "",
                        )
                    ),

                    "reported_value": (
                        _safe_float(
                            observation.get(
                                "value"
                            )
                        )
                    ),

                    "reported_unit": (
                        observation.get(
                            "unit",
                            "",
                        )
                    ),

                    "density_override_g_cm3": "",

                    "notes": (
                        observation.get(
                            "notes",
                            "",
                        )
                    ),
                }
            )

        # Blank rows for new records

        for _ in range(
            blank_rows_per_site
        ):
            rows_for_site.append(
                {
                    "source_code": source_code,
                    "source_title": source_title,

                    "site_id": site_id,
                    "site_label": site_label,
                    "country": country,

                    "observation_id": "",

                    "material": "",
                    "exposure_period": "",
                    "exposure_start": "",
                    "exposure_end": "",
                    "corrosion_metric": "",

                    "reported_value": "",
                    "reported_unit": "",

                    "density_override_g_cm3": "",

                    "notes": "",
                }
            )

        first_row_by_site[
            site_id
        ] = sheet.max_row + 1

        site_first_row = (
            sheet.max_row + 1
        )

        for record in rows_for_site:

            output_row = [
                # A — source_code
                record.get(
                    "source_code",
                    "",
                ),

                # B — source_title
                record.get(
                    "source_title",
                    "",
                ),

                # C — site_id
                record.get(
                    "site_id",
                    "",
                ),

                # D — site_label
                record.get(
                    "site_label",
                    "",
                ),

                # E — country
                record.get(
                    "country",
                    "",
                ),

                # F — observation_id
                record.get(
                    "observation_id",
                    "",
                ),

                # G — material
                record.get(
                    "material",
                    "",
                ),

                # H — exposure_period
                record.get(
                    "exposure_period",
                    "",
                ),

                # I — exposure_start
                record.get(
                    "exposure_start",
                    "",
                ),

                # J — exposure_end
                record.get(
                    "exposure_end",
                    "",
                ),

                # K — corrosion_metric
                record.get(
                    "corrosion_metric",
                    "",
                ),

                # L — reported_value
                record.get(
                    "reported_value",
                    "",
                ),

                # M — reported_unit
                record.get(
                    "reported_unit",
                    "",
                ),

                # N — default_density_g_cm3
                "",

                # O — density_override_g_cm3
                record.get(
                    "density_override_g_cm3",
                    "",
                ),

                # P — density_used_g_cm3
                "",

                # Q — canonical_thickness_loss_rate_um_year
                "",

                # R — canonical_mass_loss_rate_g_m2_year
                "",

                # S — normalization_note
                "",

                # T — notes
                record.get(
                    "notes",
                    "",
                ),
            ]

            sheet.append(output_row)

    last_data_row = sheet.max_row

    site_separator = Side(
        style="medium",
        color="4472C4",
    )

    sheet.auto_filter.ref = (
        f"A1:"
        f"{get_column_letter(len(WORKBOOK_COLUMNS))}"
        f"{last_data_row}"
    )

    if last_data_row < first_data_row:
        raise ValueError(
            "No corrosion workbook rows "
            "were generated."
        )

    lists_name = quote_sheetname(
        lists_sheet.title
    )

    # =========================================================
    # Calculated columns
    #
    # =========================================================
    # Calculated corrosion outputs
    #
    # Q = canonical general thickness-loss rate, µm/year
    # R = canonical general mass-loss rate, g/m²/year
    #
    # V and W are hidden compatibility helper columns:
    #
    # V = reported-unit conversion factor
    # W = exposure duration converted to years
    #
    # The curator app remains authoritative and independently
    # recalculates these values during import.
    # =========================================================
    #
    # Conversion paths:
    #
    # penetration_rate
    #     -> unit conversion
    #     -> µm/year
    #
    # mass_loss_rate
    #     -> g/m²/year
    #     -> divide by density
    #     -> µm/year
    #
    # cumulative_penetration
    #     -> µm
    #     -> divide by exposure duration
    #     -> µm/year
    #
    # cumulative_mass_loss
    #     -> g/m²
    #     -> divide by density
    #     -> µm
    #     -> divide by exposure duration
    #     -> µm/year
    #
    # MCI / Al-ACI / ICI / maximum pit depth /
    # net mass change are intentionally NOT converted
    # to general corrosion rate.
    # =========================================================

    for row_number in range(
        first_data_row,
        last_data_row + 1,
    ):

        sheet.row_dimensions[
            row_number
        ].height = 24

        material_cell = (
            f"G{row_number}"
        )

        exposure_cell = (
            f"H{row_number}"
        )

        metric_cell = (
            f"K{row_number}"
        )

        value_cell = (
            f"L{row_number}"
        )

        unit_cell = (
            f"M{row_number}"
        )

        default_density_cell = (
            f"N{row_number}"
        )

        density_override_cell = (
            f"O{row_number}"
        )

        density_used_cell = (
            f"P{row_number}"
        )

        thickness_rate_cell = (
            f"Q{row_number}"
        )

        mass_loss_rate_cell = (
            f"R{row_number}"
        )

        note_cell = (
            f"S{row_number}"
        )

        unit_factor_cell = (
            f"V{row_number}"
        )

        exposure_years_cell = (
            f"W{row_number}"
        )

        # -----------------------------------------------------
        # Default material density
        # -----------------------------------------------------

        sheet[
            default_density_cell
        ] = (
            f'=IF('
            f'{material_cell}="",'
            f'"",'
            f'IFERROR('
            f'VLOOKUP('
            f'{material_cell},'
            f'{lists_name}!'
            f'$E${density_start_row}:'
            f'$F${density_end_row},'
            f'2,'
            f'FALSE'
            f'),'
            f'""'
            f')'
            f')'
        )

        # -----------------------------------------------------
        # Density actually used
        # -----------------------------------------------------

        sheet[
            density_used_cell
        ] = (
            f'=IF('
            f'{density_override_cell}<>"",'
            f'{density_override_cell},'
            f'{default_density_cell}'
            f')'
        )

        # -----------------------------------------------------
        # Hidden helper: reported-unit conversion factor
        # -----------------------------------------------------

        sheet[
            unit_factor_cell
        ] = (
            f'=IF('
            f'OR('
            f'{metric_cell}="",'
            f'{unit_cell}=""'
            f'),'
            f'"",'
            f'IFERROR('
            f'VLOOKUP('
            f'{metric_cell}&"|"&{unit_cell},'
            f'{lists_name}!$A$2:$C${unit_rule_end_row},'
            f'3,'
            f'FALSE'
            f'),'
            f'""'
            f')'
            f')'
        )

        # -----------------------------------------------------
        # Hidden helper: exposure duration in years
        #
        # Supports:
        # 1 year / years / yr / yrs / y
        # 6 months / month / mo / mos
        # 3 weeks / week / wk / wks
        # 90 days / day / d
        # -----------------------------------------------------

        sheet[
            exposure_years_cell
        ] = (
            f'=IF('
            f'{exposure_cell}="",'
            f'"",'
            f'IFERROR('

            f'VALUE('
            f'LEFT('
            f'TRIM({exposure_cell}),'
            f'FIND(" ",TRIM({exposure_cell})&" ")-1'
            f')'
            f')*'

            f'IF('
            f'OR('
            f'ISNUMBER(SEARCH("year",LOWER({exposure_cell}))),'
            f'ISNUMBER(SEARCH(" yr",LOWER({exposure_cell}))),'
            f'RIGHT(LOWER(TRIM({exposure_cell})),1)="y"'
            f'),'
            f'1,'

            f'IF('
            f'OR('
            f'ISNUMBER(SEARCH("month",LOWER({exposure_cell}))),'
            f'ISNUMBER(SEARCH(" mo",LOWER({exposure_cell})))'
            f'),'
            f'1/12,'

            f'IF('
            f'OR('
            f'ISNUMBER(SEARCH("week",LOWER({exposure_cell}))),'
            f'ISNUMBER(SEARCH(" wk",LOWER({exposure_cell})))'
            f'),'
            f'7/365.25,'

            f'IF('
            f'OR('
            f'ISNUMBER(SEARCH("day",LOWER({exposure_cell}))),'
            f'RIGHT(LOWER(TRIM({exposure_cell})),1)="d"'
            f'),'
            f'1/365.25,'
            f'""'
            f')'
            f')'
            f')'
            f'),'

            f'""'
            f')'
            f')'
        )

        # -----------------------------------------------------
        # FINAL canonical corrosion rate
        #
        # Exposure duration parser accepts entries such as:
        #
        # 1 year
        # 2 years
        # 0.5 year
        # 6 months
        # 18 months
        # 90 days
        # 3 weeks
        #
        # The Python importer still performs the authoritative
        # calculation independently.
        # -----------------------------------------------------

        # -----------------------------------------------------
        # Canonical thickness-loss rate — µm/year
        # -----------------------------------------------------

        sheet[
            thickness_rate_cell
        ] = (
            f'=IF('
            f'OR('
            f'{metric_cell}="",'
            f'{value_cell}="",'
            f'{unit_cell}=""'
            f'),'
            f'"",'

            # Direct penetration-rate observation
            f'IF('
            f'OR('
            f'{metric_cell}="penetration_rate",'
            f'{metric_cell}="corrosion_rate"'
            f'),'
            f'IF('
            f'{unit_factor_cell}="",'
            f'"",'
            f'{value_cell}*{unit_factor_cell}'
            f'),'

            # Mass-loss rate -> penetration rate
            f'IF('
            f'{metric_cell}="mass_loss_rate",'
            f'IF('
            f'OR('
            f'{unit_factor_cell}="",'
            f'{density_used_cell}=""'
            f'),'
            f'"",'
            f'{value_cell}*{unit_factor_cell}/{density_used_cell}'
            f'),'

            # Cumulative penetration -> average annual rate
            f'IF('
            f'{metric_cell}="cumulative_penetration",'
            f'IF('
            f'OR('
            f'{unit_factor_cell}="",'
            f'{exposure_years_cell}=""'
            f'),'
            f'"",'
            f'{value_cell}*{unit_factor_cell}/{exposure_years_cell}'
            f'),'

            # Cumulative mass loss -> average penetration rate
            f'IF('
            f'{metric_cell}="cumulative_mass_loss",'
            f'IF('
            f'OR('
            f'{unit_factor_cell}="",'
            f'{density_used_cell}="",'
            f'{exposure_years_cell}=""'
            f'),'
            f'"",'
            f'{value_cell}*{unit_factor_cell}/'
            f'{density_used_cell}/'
            f'{exposure_years_cell}'
            f'),'

            # Non-general-corrosion metric
            f'""'

            f')'
            f')'
            f')'
            f')'
            f')'
        )

        # -----------------------------------------------------
        # Canonical mass-loss rate — g/m²/year
        # -----------------------------------------------------

        sheet[
            mass_loss_rate_cell
        ] = (
            f'=IF('
            f'OR('
            f'{metric_cell}="",'
            f'{value_cell}="",'
            f'{unit_cell}=""'
            f'),'
            f'"",'

            # Penetration rate -> mass-loss rate
            f'IF('
            f'OR('
            f'{metric_cell}="penetration_rate",'
            f'{metric_cell}="corrosion_rate"'
            f'),'
            f'IF('
            f'OR('
            f'{unit_factor_cell}="",'
            f'{density_used_cell}=""'
            f'),'
            f'"",'
            f'{value_cell}*{unit_factor_cell}*{density_used_cell}'
            f'),'

            # Direct mass-loss-rate observation
            f'IF('
            f'{metric_cell}="mass_loss_rate",'
            f'IF('
            f'{unit_factor_cell}="",'
            f'"",'
            f'{value_cell}*{unit_factor_cell}'
            f'),'

            # Cumulative penetration -> annual mass loss
            f'IF('
            f'{metric_cell}="cumulative_penetration",'
            f'IF('
            f'OR('
            f'{unit_factor_cell}="",'
            f'{density_used_cell}="",'
            f'{exposure_years_cell}=""'
            f'),'
            f'"",'
            f'{value_cell}*{unit_factor_cell}*'
            f'{density_used_cell}/'
            f'{exposure_years_cell}'
            f'),'

            # Cumulative mass loss -> annual mass-loss rate
            f'IF('
            f'{metric_cell}="cumulative_mass_loss",'
            f'IF('
            f'OR('
            f'{unit_factor_cell}="",'
            f'{exposure_years_cell}=""'
            f'),'
            f'"",'
            f'{value_cell}*{unit_factor_cell}/'
            f'{exposure_years_cell}'
            f'),'

            # Non-general-corrosion metric
            f'""'

            f')'
            f')'
            f')'
            f')'
            f')'
        )

        # -----------------------------------------------------
        # Human-readable conversion note
        # -----------------------------------------------------

        note_formula = (
            f'=IF('
            f'OR('
            f'{metric_cell}="",'
            f'{value_cell}="",'
            f'{unit_cell}=""'
            f'),'
            f'"",'

            f'IF('
            f'OR('
            f'{metric_cell}="maximum_pit_depth",'
            f'{metric_cell}="net_mass_change",'
            f'{metric_cell}="MCI",'
            f'{metric_cell}="Al-ACI",'
            f'{metric_cell}="ICI"'
            f'),'
            f'"Not converted to canonical general-corrosion rates",'

            f'IF('
            f'AND('
            f'{exposure_cell}="",'
            f'OR('
            f'{metric_cell}="cumulative_penetration",'
            f'{metric_cell}="cumulative_mass_loss"'
            f')'
            f'),'
            f'"Exposure period not reported — cumulative value preserved; annual canonical rates unavailable",'            

            f'IF('
            f'AND('
            f'{thickness_rate_cell}<>"",'
            f'{mass_loss_rate_cell}<>""'
            f'),'
            f'"Both canonical rates calculated",'

            f'IF('
            f'{thickness_rate_cell}<>"",'
            f'"Thickness-loss rate calculated; mass-loss rate unavailable",'

            f'IF('
            f'{mass_loss_rate_cell}<>"",'
            f'"Mass-loss rate calculated; thickness-loss rate unavailable",'

            f'"Canonical rates not calculated — check metric, unit, exposure duration, or density"'
            f')'
            f')'
            f')'
            f')'
            f')'
            f')'
        )

        if (
            note_formula.count("(")
            != note_formula.count(")")
        ):
            raise RuntimeError(
                "Generated normalization-note formula "
                f"is unbalanced at {note_cell}: "
                f"{note_formula.count('(')} opening "
                f"parentheses vs "
                f"{note_formula.count(')')} closing."
            )

        sheet[
            note_cell
        ] = note_formula

    # =========================================================
    # Data validation
    # =========================================================

    metal_validation = DataValidation(
        type="list",
        formula1="=CorrosionMetalOptions",
        allow_blank=True,
    )

    metal_validation.error = (
        "Choose a material from the curator list."
    )

    metal_validation.errorTitle = (
        "Invalid material"
    )

    metal_validation.showErrorMessage = True

    metal_validation.showInputMessage = True

    metal_validation.promptTitle = (
        "Material"
    )

    metal_validation.prompt = (
        "Choose a material already registered "
        "in the curator database."
    )

    # Exposure values can be manually overridden/customized.

    exposure_validation = DataValidation(
        type="list",
        formula1="=CorrosionExposureOptions",
        allow_blank=True,
    )

    exposure_validation.showErrorMessage = False

    exposure_validation.showInputMessage = True

    exposure_validation.promptTitle = (
        "Exposure duration"
    )

    exposure_validation.prompt = (
        "Optional. Choose a known duration, type a custom "
        "duration, or leave blank if the source does not "
        "report the exposure period."
    )

    metric_validation = DataValidation(
        type="list",
        formula1="=CorrosionMetricOptions",
        allow_blank=True,
    )

    metric_validation.error = (
        "Choose an approved corrosion metric."
    )

    metric_validation.errorTitle = (
        "Invalid metric"
    )

    metric_validation.showErrorMessage = True

    metric_validation.showInputMessage = True

    metric_validation.promptTitle = (
        "Reported corrosion metric"
    )

    metric_validation.prompt = (
        "Choose the physical quantity reported by the source. "
        "Do not convert the source value manually."
    )

    unit_validation = DataValidation(
        type="list",
        formula1=(
            '=INDIRECT('
            'IF('
            '$K2="",'
            '"CorrosionUnits_blank",'
            '"CorrosionUnits_"&'
            'SUBSTITUTE($K2,"-","_")'
            ')'
            ')'
        ),
        allow_blank=True,
    )

    unit_validation.error = (
        "Choose one of the approved reported units."
    )

    unit_validation.errorTitle = (
        "Invalid unit"
    )

    unit_validation.showErrorMessage = True

    unit_validation.showInputMessage = True

    unit_validation.promptTitle = (
        "Reported unit"
    )

    unit_validation.prompt = (
        "Choose the unit exactly as reported by the source. "
        "The workbook will calculate canonical rates automatically."
    )

    numeric_value_validation = DataValidation(
        type="decimal",
        operator="between",
        formula1="-1E+100",
        formula2="1E+100",
        allow_blank=True,
    )

    numeric_value_validation.error = (
        "Enter a numerical reported value."
    )

    numeric_value_validation.errorTitle = (
        "Invalid number"
    )

    numeric_value_validation.showErrorMessage = True

    density_validation = DataValidation(
        type="decimal",
        operator="greaterThan",
        formula1="0",
        allow_blank=True,
    )

    density_validation.error = (
        "Density must be greater than zero."
    )

    density_validation.errorTitle = (
        "Invalid density"
    )

    density_validation.showErrorMessage = True

    for validation in [
        metal_validation,
        exposure_validation,
        metric_validation,
        unit_validation,
        numeric_value_validation,
        density_validation,
    ]:
        sheet.add_data_validation(
            validation
        )

    metal_validation.add(
        f"G{first_data_row}:"
        f"G{last_data_row}"
    )

    exposure_validation.add(
        f"H{first_data_row}:"
        f"H{last_data_row}"
    )

    metric_validation.add(
        f"K{first_data_row}:"
        f"K{last_data_row}"
    )

    numeric_value_validation.add(
        f"L{first_data_row}:"
        f"L{last_data_row}"
    )

    unit_validation.add(
        f"M{first_data_row}:"
        f"M{last_data_row}"
    )

    density_validation.add(
        f"O{first_data_row}:"
        f"O{last_data_row}"
    )

    # =========================================================
    # Formatting
    # =========================================================

    identity_fill = PatternFill(
        "solid",
        fgColor="EAF2F8",
    )

    entry_fill = PatternFill(
        "solid",
        fgColor="FFF2CC",
    )

    calculated_fill = PatternFill(
        "solid",
        fgColor="E7E6E6",
    )

    existing_fill = PatternFill(
        "solid",
        fgColor="E2F0D9",
    )

    warning_fill = PatternFill(
        "solid",
        fgColor="FCE4D6",
    )

    incomplete_fill = PatternFill(
        "solid",
        fgColor="F4CCCC",
    )

    for row_number in range(
        first_data_row,
        last_data_row + 1,
    ):

        # Source/site/ID context

        for column_index in range(
            1,
            7,
        ):
            sheet.cell(
                row=row_number,
                column=column_index,
            ).fill = identity_fill

        # Human-entry fields

        for column_index in [
            7,   # material
            8,   # exposure period
            9,   # exposure start
            10,  # exposure end
            11,  # metric
            12,  # reported value
            13,  # reported unit
            15,  # density override
            20,  # notes
        ]:
            sheet.cell(
                row=row_number,
                column=column_index,
            ).fill = entry_fill

        # Calculated columns

        for column_index in [
            14,  # default density
            16,  # density used
            17,  # canonical thickness-loss rate
            18,  # canonical mass-loss rate
            19,  # normalization note
        ]:
            cell = sheet.cell(
                row=row_number,
                column=column_index,
            )

            cell.fill = calculated_fill

            cell.protection = Protection(
                locked=True
            )

        # Canonical outputs are the primary derived quantities.

        for column_index in [
            17,
            18,
        ]:
            sheet.cell(
                row=row_number,
                column=column_index,
            ).font = Font(
                bold=True,
                color="375623",
            )

        # Existing database rows are subtly green.

        observation_id = sheet.cell(
            row=row_number,
            column=6,
        ).value

        if observation_id not in (
            None,
            "",
        ):
            for column_index in range(
                1,
                len(WORKBOOK_COLUMNS) + 1,
            ):
                # Keep calculated fields grey.
                if column_index not in [
                    14,
                    16,
                    17,
                    18,
                    19,
                ]:
                    sheet.cell(
                        row=row_number,
                        column=column_index,
                    ).fill = existing_fill

                    sheet.cell(
                        row=row_number,
                        column=6,
                    ).font = Font(
                        bold=True,
                        color="375623",
                    )

        for column_index in range(
            1,
            len(WORKBOOK_COLUMNS) + 1,
        ):
            cell = sheet.cell(
                row=row_number,
                column=column_index,
            )

            cell.border = Border(
                bottom=thin_gray
            )

            cell.alignment = Alignment(
                horizontal=(
                    "left"
                    if column_index in [
                        6,
                        7,
                    ]
                    else (
                        "right"
                        if column_index in [
                            12,  # reported value
                            14,  # default density
                            15,  # density override
                            16,  # density used
                            17,  # canonical thickness loss
                            18,  # canonical mass loss
                        ]
                        else None
                    )
                ),

                vertical="top",
                wrap_text=(
                    column_index
                    in [
                        2,
                        4,
                        19,
                        20,
                    ]
                ),
            )


    # ---------------------------------------------------------
    # Site block separators
    # ---------------------------------------------------------

    for site_first_row in (
        first_row_by_site.values()
    ):
        for column_index in range(
            1,
            len(WORKBOOK_COLUMNS) + 1,
        ):
            cell = sheet.cell(
                row=site_first_row,
                column=column_index,
            )

            cell.border = Border(
                top=site_separator,
                bottom=thin_gray,
            )

        for column_index in range(
            1,
            7,
        ):
            sheet.cell(
                row=site_first_row,
                column=column_index,
            ).font = Font(
                bold=True,
                color="17365D",
            )

    # Leave visual room at the left edge of material cells
    # for the VBA + row-duplication button.
    for row_number in range(
        first_data_row,
        last_data_row + 1,
    ):
        sheet.cell(
            row=row_number,
            column=7,
        ).alignment = Alignment(
            horizontal="left",
            vertical="top",
            indent=4,
        )

    # ---------------------------------------------------------
    # Incomplete observation rows
    #
    # If at least one observation-input field G:M has been entered
    # but one or more required fields are still blank, highlight
    # the input area in pale red.
    # ---------------------------------------------------------

    sheet.conditional_formatting.add(
        f"G{first_data_row}:"
        f"M{last_data_row}",

        FormulaRule(
            formula=[
                f'AND('
                f'COUNTA($G{first_data_row}:$M{first_data_row})>0,'
                f'OR('
                f'$G{first_data_row}="",'
                f'$K{first_data_row}="",'
                f'$L{first_data_row}="",'
                f'$M{first_data_row}=""'
                f')'
                f')'
            ],

            fill=incomplete_fill,
        ),
    )

    # Unsupported metric/unit combinations become orange.

    sheet.conditional_formatting.add(
        f"S{first_data_row}:"
        f"S{last_data_row}",

        FormulaRule(
            formula=[
                f'ISNUMBER('
                f'SEARCH('
                f'"Unsupported",'
                f'S{first_data_row}'
                f')'
                f')'
            ],

            fill=warning_fill,
        ),
    )

    sheet.conditional_formatting.add(
        f"S{first_data_row}:"
        f"S{last_data_row}",

        FormulaRule(
            formula=[
                f'OR('
                f'ISNUMBER(SEARCH("unavailable",S{first_data_row})),'
                f'ISNUMBER(SEARCH("not calculated",S{first_data_row}))'
                f')'
            ],

            fill=warning_fill,
        ),
    )    

    # Exposure calendar dates are deliberately stored as text.
    # This preserves partial source precision such as:
    #
    # 1994
    # 1994-12
    # 1994-12-15
    #
    # It also lets the VBA normalizer see natural-language
    # curator input before Excel converts it to a serial date.

    for row_number in range(
        first_data_row,
        last_data_row + 1,
    ):
        sheet.cell(
            row=row_number,
            column=9,   # exposure_start
        ).number_format = "@"

        sheet.cell(
            row=row_number,
            column=10,  # exposure_end
        ).number_format = "@"

    # Number display

    for row_number in range(
        first_data_row,
        last_data_row + 1,
    ):

        for column_index in [
            12,  # reported value
            14,  # default density
            15,  # density override
            16,  # density used
            17,  # canonical thickness-loss rate
            18,  # canonical mass-loss rate
        ]:
            sheet.cell(
                row=row_number,
                column=column_index,
            ).number_format = (
                "0.0000############"
            )

        for column_index in [
            17,
            18,
        ]:
            sheet.cell(
                row=row_number,
                column=column_index,
            ).number_format = (
                "0.000"
            )

    # =========================================================
    # Column widths
    # =========================================================

    column_widths = {
        "A": 12,
        "B": 30,
        "C": 12,
        "D": 24,
        "E": 18,
        "F": 18,

        "G": 24,
        "H": 17,
        "I": 16,
        "J": 16,
        "K": 25,

        "L": 15,
        "M": 18,

        "N": 18,
        "O": 19,
        "P": 22,

        "Q": 29,
        "R": 29,

        "S": 46,
        "T": 36,
    }

    for (
        column_letter,
        width,
    ) in column_widths.items():

        sheet.column_dimensions[
            column_letter
        ].width = width

    # =========================================================
    # Final worksheet visibility / view reset
    # =========================================================
    #
    # Normal curator-facing columns:
    #
    #   A  source_code
    #   C  site_id
    #   D  site_label
    #   E  country
    #   F  observation_id
    #   G  material
    #   H  exposure_period
    #   I  exposure_start
    #   J  exposure_end
    #   K  corrosion_metric
    #   L  reported_value
    #   M  reported_unit
    #   P  density_used_g_cm3
    #   Q  canonical thickness-loss rate
    #   R  canonical mass-loss rate
    #   S  normalization_note
    #   T  notes
    #
    #   Hidden:
    #
    #    B     source_title
    #    N:O   density details
    #    U:W   workbook-internal columns
    #
    # There are NO native Excel outline groups.
    # =========================================================

    # ---------------------------------------------------------
    # Remove all stale outline/group state
    # ---------------------------------------------------------

    sheet.sheet_view.showOutlineSymbols = False

    sheet.sheet_format.outlineLevelCol = 0
    sheet.sheet_format.outlineLevelRow = 0

    for column_letter in [
        "A",
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "H",
        "I",
        "J",
        "K",
        "L",
        "M",
        "N",
        "O",
        "P",
        "Q",
        "R",
        "S",
        "T",
        "U",
        "V",
        "W",
    ]:
        dimension = sheet.column_dimensions[
            column_letter
        ]

        dimension.outlineLevel = 0
        dimension.collapsed = False

        # Start from a completely visible state.
        dimension.hidden = False

    # ---------------------------------------------------------
    # Permanently hidden source title
    # ---------------------------------------------------------

    sheet.column_dimensions[
        "B"
    ].hidden = True

    # ---------------------------------------------------------
    # Density details
    #
    # These are hidden initially and controlled by the
    # dedicated VBA density +/- button.
    # ---------------------------------------------------------

    sheet.column_dimensions[
        "N"
    ].hidden = True

    sheet.column_dimensions[
        "O"
    ].hidden = True

    # P = density_used_g_cm3 must remain visible.
    sheet.column_dimensions[
        "P"
    ].hidden = False

    # ---------------------------------------------------------
    # Internal workbook columns
    # ---------------------------------------------------------

    sheet.column_dimensions[
        "U"
    ].hidden = True

    sheet.column_dimensions[
        "V"
    ].hidden = True

    sheet.column_dimensions[
        "W"
    ].hidden = True

    # ---------------------------------------------------------
    # Explicitly guarantee all normal curator columns are visible
    # ---------------------------------------------------------

    for column_letter in [
        "A",
        "C",
        "D",
        "E",
        "F",
        "G",
        "H",
        "I",
        "J",
        "K",
        "L",
        "M",
        "P",
        "Q",
        "R",
        "S",
        "T",
    ]:
        sheet.column_dimensions[
            column_letter
        ].hidden = False

    # ---------------------------------------------------------
    # Workbook appearance
    # ---------------------------------------------------------

    sheet.sheet_view.showGridLines = False

    sheet.sheet_properties.tabColor = "17365D"

    sheet.sheet_view.zoomScale = 90
    sheet.sheet_view.zoomScaleNormal = 90

    # ---------------------------------------------------------
    # CRITICAL: reset the saved Excel viewport
    #
    # The reusable XLSM template may have been saved while
    # horizontally scrolled far to the right. Without this,
    # a newly generated workbook can open around column K/AA
    # even though columns A:J still exist.
    # ---------------------------------------------------------

    sheet.sheet_view.topLeftCell = "A1"

    for selection in sheet.sheet_view.selection:
        selection.activeCell = "A1"
        selection.sqref = "A1"

    workbook.active = workbook.sheetnames.index(
        "Corrosion Observations"
    )

    # ---------------------------------------------------------
    # Force Excel to recalculate workbook formulas
    # ---------------------------------------------------------

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcOnSave = True

    # ---------------------------------------------------------
    # Save
    # ---------------------------------------------------------

    output = BytesIO()

    workbook.save(output)

    return output.getvalue()

def build_corrosion_entry_workbook_xlsm(
    *,
    source_row: dict[str, Any],
    site_links: list[dict[str, Any]],
    site_lookup: dict[str, dict[str, Any]],
    existing_observations: list[dict[str, Any]],
    metal_options: list[str],
    exposure_options: list[str],
) -> bytes:
    """
    Build the macro-enabled corrosion entry workbook.

    The XLSM version contains one starter row per linked
    site. Additional rows are created directly in Excel
    using the + action column.
    """

    if not CORROSION_XLSM_TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            "The macro-enabled workbook template is missing. "
            "Expected file: "
            f"{CORROSION_XLSM_TEMPLATE_PATH.name}"
        )

    return build_corrosion_entry_workbook(
        source_row=source_row,
        site_links=site_links,
        site_lookup=site_lookup,
        existing_observations=(
            existing_observations
        ),
        metal_options=metal_options,
        exposure_options=(
            exposure_options
        ),
        blank_rows_per_site=1,
        macro_template_path=(
            CORROSION_XLSM_TEMPLATE_PATH
        ),
    )

def read_corrosion_entry_workbook(
    uploaded_file,
) -> pd.DataFrame:
    """
    Read a curator-generated corrosion workbook.

    Calculated Excel columns are intentionally ignored.
    The app recalculates normalization itself.

    Completely blank observation rows are ignored.
    """

    uploaded_file.seek(0)

    workbook = load_workbook(
        uploaded_file,
        read_only=True,
        data_only=False,
    )

    if "Corrosion Observations" not in workbook.sheetnames:
        raise ValueError(
            "Workbook does not contain the required "
            "`Corrosion Observations` sheet."
        )

    sheet = workbook[
        "Corrosion Observations"
    ]

    raw_rows = list(
        sheet.iter_rows(
            values_only=True
        )
    )

    if not raw_rows:
        raise ValueError(
            "The corrosion workbook is empty."
        )

    headers = [
        str(value or "").strip()
        for value in raw_rows[0]
    ]

    missing_columns = [
        column
        for column in WORKBOOK_INPUT_COLUMNS
        if column not in headers
    ]

    if missing_columns:
        raise ValueError(
            "The workbook is missing required column(s): "
            + ", ".join(missing_columns)
        )

    row_records: list[
        dict[str, Any]
    ] = []

    for excel_row_number, values in enumerate(
        raw_rows[1:],
        start=2,
    ):
        raw_record = {
            headers[index]: (
                values[index]
                if index < len(values)
                else ""
            )
            for index in range(
                len(headers)
            )
        }

        # Ignore the workbook formula/calculation outputs.
        # Python will regenerate all of them.

        record = {
            "excel_row": excel_row_number,

            "source_code": str(
                raw_record.get(
                    "source_code",
                    "",
                ) or ""
            ).strip(),

            "source_title": str(
                raw_record.get(
                    "source_title",
                    "",
                ) or ""
            ).strip(),

            "site_id": str(
                raw_record.get(
                    "site_id",
                    "",
                ) or ""
            ).strip(),

            "site_label": str(
                raw_record.get(
                    "site_label",
                    "",
                ) or ""
            ).strip(),

            "country": str(
                raw_record.get(
                    "country",
                    "",
                ) or ""
            ).strip(),

            "observation_id": (
                raw_record.get(
                    "observation_id",
                    "",
                )
            ),

            "material": str(
                raw_record.get(
                    "material",
                    "",
                ) or ""
            ).strip(),

            "exposure_period": str(
                raw_record.get(
                    "exposure_period",
                    "",
                ) or ""
            ).strip(),

            "exposure_start": str(
                raw_record.get(
                    "exposure_start",
                    "",
                ) or ""
            ).strip(),

            "exposure_end": str(
                raw_record.get(
                    "exposure_end",
                    "",
                ) or ""
            ).strip(),

            "corrosion_metric": str(
                raw_record.get(
                    "corrosion_metric",
                    "",
                ) or ""
            ).strip(),

            "reported_value": (
                raw_record.get(
                    "reported_value",
                    "",
                )
            ),

            "reported_unit": (
                _normalize_unit_text(
                    raw_record.get(
                        "reported_unit",
                        "",
                    )
                )
            ),

            "density_override_g_cm3": (
                raw_record.get(
                    "density_override_g_cm3",
                    "",
                )
            ),

            "notes": str(
                raw_record.get(
                    "notes",
                    "",
                ) or ""
            ).strip(),
        }

        # A generated source/site row contains context even when
        # no observation has been entered. Determine whether any
        # observation-entry field is actually populated.

        observation_entry_values = [
            record.get(
                "observation_id",
                ""
            ),
            record.get(
                "material",
                ""
            ),
            record.get(
                "exposure_period",
                ""
            ),
            record.get(
                "exposure_start",
                ""
            ),
            record.get(
                "exposure_end",
                ""
            ),
            record.get(
                "corrosion_metric",
                ""
            ),
            record.get(
                "reported_value",
                ""
            ),
            record.get(
                "reported_unit",
                ""
            ),
            record.get(
                "density_override_g_cm3",
                ""
            ),
            record.get(
                "notes",
                ""
            ),
        ]

        has_observation_content = any(
            str(value or "").strip()
            for value
            in observation_entry_values
        )

        if not has_observation_content:
            continue

        row_records.append(
            record
        )

    return pd.DataFrame(
        row_records
    )

def validate_corrosion_workbook_rows(
    dataframe: pd.DataFrame,
    *,
    existing_site_ids: set[str],
    existing_source_codes: set[str],
    existing_site_source_pairs: set[
        tuple[str, str]
    ],
    existing_observations: list[
        dict[str, Any]
    ],
) -> pd.DataFrame:
    """
    Validate workbook rows before database import.

    This function does not write anything to the database.
    """

    if dataframe.empty:
        return dataframe.copy()

    preview_df = dataframe.copy()

    preview_columns = [
        "record_action",
        "validation_status",
        "validation_message",

        "canonical_thickness_loss_rate_um_year",
        "canonical_mass_loss_rate_g_m2_year",

        "default_density_g_cm3",
        "density_used_g_cm3",
        "density_basis",

        "normalization_note",
    ]

    for column in preview_columns:
        if column not in preview_df.columns:
            preview_df[column] = pd.Series(
                "",
                index=preview_df.index,
                dtype="object",
            )
        else:
            preview_df[column] = (
                preview_df[column]
                .astype("object")
            )

    normalized_existing_sites = {
        str(value).strip().casefold()
        for value
        in existing_site_ids
        if str(value).strip()
    }

    normalized_existing_sources = {
        str(value).strip().casefold()
        for value
        in existing_source_codes
        if str(value).strip()
    }

    normalized_site_source_pairs = {
        (
            str(site_id).strip().casefold(),
            str(source_code).strip().casefold(),
        )
        for (
            site_id,
            source_code,
        )
        in existing_site_source_pairs
    }

    observations_by_id: dict[
        int,
        dict[str, Any]
    ] = {}

    for observation in existing_observations:
        try:
            observation_id = int(
                observation.get(
                    "id"
                )
            )
        except (
            TypeError,
            ValueError,
        ):
            continue

        observations_by_id[
            observation_id
        ] = observation

    for row_index, row in preview_df.iterrows():
        errors: list[str] = []

        site_id = str(
            row.get(
                "site_id",
                "",
            ) or ""
        ).strip()

        source_code = str(
            row.get(
                "source_code",
                "",
            ) or ""
        ).strip()

        material = str(
            row.get(
                "material",
                "",
            ) or ""
        ).strip()

        exposure_period = str(
            row.get(
                "exposure_period",
                "",
            ) or ""
        ).strip()

        exposure_start = str(
            row.get(
                "exposure_start",
                "",
            ) or ""
        ).strip()

        exposure_end = str(
            row.get(
                "exposure_end",
                "",
            ) or ""
        ).strip()

        if (
            exposure_start
            and not is_valid_partial_iso_date(
                exposure_start
            )
        ):
            errors.append(
                "exposure_start must use YYYY, "
                "YYYY-MM, or YYYY-MM-DD."
            )

        if (
            exposure_end
            and not is_valid_partial_iso_date(
                exposure_end
            )
        ):
            errors.append(
                "exposure_end must use YYYY, "
                "YYYY-MM, or YYYY-MM-DD."
            )

        corrosion_metric = str(
            row.get(
                "corrosion_metric",
                "",
            ) or ""
        ).strip()

        reported_unit = (
            _normalize_unit_text(
                row.get(
                    "reported_unit",
                    "",
                )
            )
        )

        observation_id_raw = (
            row.get(
                "observation_id",
                ""
            )
        )

        observation_id: int | None = None

        if str(
            observation_id_raw or ""
        ).strip():
            try:
                observation_id_number = float(
                    observation_id_raw
                )

                if not (
                    observation_id_number.is_integer()
                ):
                    raise ValueError

                observation_id = int(
                    observation_id_number
                )
            except (
                TypeError,
                ValueError,
            ):
                errors.append(
                    "observation_id is not a valid integer."
                )

        # ---------------------------------------------
        # Required fields
        # ---------------------------------------------

        required_values = {
            "source_code": source_code,
            "site_id": site_id,
            "material": material,
            "corrosion_metric": corrosion_metric,
            "reported_value": row.get(
                "reported_value",
                "",
            ),
            "reported_unit": reported_unit,
        }

        for (
            field_name,
            field_value,
        ) in required_values.items():

            if str(
                field_value
                if field_value is not None
                else ""
            ).strip() == "":
                errors.append(
                    f"{field_name} is required."
                )

        # ---------------------------------------------
        # Site / source / site-source provenance
        # ---------------------------------------------

        site_key = site_id.casefold()
        source_key = source_code.casefold()

        if (
            site_id
            and site_key
            not in normalized_existing_sites
        ):
            errors.append(
                f"site_id `{site_id}` does not exist."
            )

        if (
            source_code
            and source_key
            not in normalized_existing_sources
        ):
            errors.append(
                f"source_code `{source_code}` does not exist."
            )

        if (
            site_id
            and source_code
            and (
                site_key,
                source_key,
            )
            not in normalized_site_source_pairs
        ):
            errors.append(
                "The selected source is not linked to "
                f"site `{site_id}`."
            )

        # ---------------------------------------------
        # Existing observation ID protection
        # ---------------------------------------------

        if observation_id is not None:
            existing_observation = (
                observations_by_id.get(
                    observation_id
                )
            )

            if existing_observation is None:
                errors.append(
                    f"observation_id `{observation_id}` "
                    "does not exist."
                )

            else:
                existing_site = str(
                    existing_observation.get(
                        "site_id",
                        "",
                    ) or ""
                ).strip()

                existing_source = str(
                    existing_observation.get(
                        "source_code",
                        "",
                    ) or ""
                ).strip()

                if (
                    existing_site.casefold()
                    != site_key
                    or
                    existing_source.casefold()
                    != source_key
                ):
                    errors.append(
                        "observation_id does not belong "
                        "to this source/site pair."
                    )

            preview_df.at[
                row_index,
                "record_action",
            ] = "UPDATE"

        else:
            preview_df.at[
                row_index,
                "record_action",
            ] = "CREATE"

        # ---------------------------------------------
        # Metric
        # ---------------------------------------------

        if (
            corrosion_metric
            and corrosion_metric
            not in CORROSION_METRIC_OPTIONS
        ):
            errors.append(
                "Unsupported corrosion_metric: "
                f"`{corrosion_metric}`."
            )

        # ---------------------------------------------
        # Numerical value
        # ---------------------------------------------

        try:
            reported_value = float(
                row.get(
                    "reported_value"
                )
            )

        except (
            TypeError,
            ValueError,
        ):
            reported_value = None

            if str(
                row.get(
                    "reported_value",
                    "",
                ) or ""
            ).strip():
                errors.append(
                    "reported_value is not numeric."
                )

        # ---------------------------------------------
        # Density override
        # ---------------------------------------------

        density_override_raw = (
            row.get(
                "density_override_g_cm3",
                "",
            )
        )

        density_override = None

        if str(
            density_override_raw or ""
        ).strip():
            try:
                density_override = float(
                    density_override_raw
                )

                if density_override <= 0:
                    errors.append(
                        "density_override_g_cm3 must "
                        "be greater than zero."
                    )

            except (
                TypeError,
                ValueError,
            ):
                errors.append(
                    "density_override_g_cm3 is not numeric."
                )

        # ---------------------------------------------
        # Recalculate normalized values
        # ---------------------------------------------

        if (
            reported_value is not None
            and corrosion_metric
            and reported_unit
        ):
            try:
                normalized = (
                    normalize_corrosion_observation(
                        material=material,
                        exposure_period=(
                            exposure_period
                        ),
                        corrosion_metric=(
                            corrosion_metric
                        ),
                        reported_value=(
                            reported_value
                        ),
                        reported_unit=(
                            reported_unit
                        ),
                        density_override_g_cm3=(
                            density_override
                        ),
                    )
                )

                preview_df.at[
                    row_index,
                    "canonical_thickness_loss_rate_um_year",
                ] = (
                    normalized[
                        "canonical_thickness_loss_rate_um_year"
                    ]
                    if normalized[
                        "canonical_thickness_loss_rate_um_year"
                    ] is not None
                    else ""
                )

                preview_df.at[
                    row_index,
                    "canonical_mass_loss_rate_g_m2_year",
                ] = (
                    normalized[
                        "canonical_mass_loss_rate_g_m2_year"
                    ]
                    if normalized[
                        "canonical_mass_loss_rate_g_m2_year"
                    ] is not None
                    else ""
                )

                preview_df.at[
                    row_index,
                    "default_density_g_cm3",
                ] = (
                    normalized[
                        "default_density_g_cm3"
                    ]
                    if normalized[
                        "default_density_g_cm3"
                    ] is not None
                    else ""
                )

                preview_df.at[
                    row_index,
                    "density_used_g_cm3",
                ] = (
                    normalized[
                        "density_g_cm3"
                    ]
                    if normalized[
                        "density_g_cm3"
                    ] is not None
                    else ""
                )

                preview_df.at[
                    row_index,
                    "density_basis",
                ] = normalized[
                    "density_basis"
                ]

                preview_df.at[
                    row_index,
                    "normalization_note",
                ] = normalized[
                    "normalization_note"
                ]

            except ValueError as exc:
                errors.append(
                    str(exc)
                )

        if errors:
            preview_df.at[
                row_index,
                "validation_status",
            ] = "ERROR"

            preview_df.at[
                row_index,
                "validation_message",
            ] = "; ".join(
                list(
                    dict.fromkeys(
                        errors
                    )
                )
            )

        else:
            preview_df.at[
                row_index,
                "validation_status",
            ] = "READY"

            preview_df.at[
                row_index,
                "validation_message",
            ] = ""

    return preview_df